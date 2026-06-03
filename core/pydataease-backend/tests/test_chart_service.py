from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from typing import Any, cast
from unittest.mock import AsyncMock
import base64

import pytest
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.schemas.auth import TokenUser
from app.services.chart_service import ChartService


@pytest.mark.asyncio
async def test_list_fields_by_dq_splits_dimension_and_quota_lists() -> None:
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))
    service.group_repo.get_by_id = AsyncMock(return_value=SimpleNamespace(id=99))  # type: ignore[attr-defined]
    service.field_repo.list_by_chart = AsyncMock(  # type: ignore[attr-defined]
        return_value=[
            SimpleNamespace(
                id=1,
                datasource_id=10,
                dataset_table_id=20,
                dataset_group_id=99,
                origin_name="region",
                name="区域",
                dataease_name="f_region",
                group_type="d",
                type="VARCHAR",
                de_type=0,
                de_extract_type=0,
                ext_field=0,
                checked=True,
                field_short_name="region",
                desensitized=False,
                params=None,
            ),
            SimpleNamespace(
                id=2,
                datasource_id=10,
                dataset_table_id=20,
                dataset_group_id=99,
                origin_name="amount",
                name="销售额",
                dataease_name="f_amount",
                group_type="q",
                type="DECIMAL",
                de_type=2,
                de_extract_type=2,
                ext_field=0,
                checked=True,
                field_short_name="amount",
                desensitized=False,
                params=None,
            ),
        ]
    )

    result = await service.list_fields_by_dq(99, 1001)

    assert result["dimensionList"][0]["groupType"] == "d"
    assert result["quotaList"][0]["groupType"] == "q"
    assert result["dimensionList"][0]["name"] == "区域"
    assert result["quotaList"][0]["name"] == "销售额"


@pytest.mark.asyncio
async def test_export_details_returns_success_stub() -> None:
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))
    service.export_task_repo.create = AsyncMock(return_value=SimpleNamespace(id="task-1"))  # type: ignore[attr-defined]
    user = TokenUser(user_id=7, oid=1)

    result = await service.export_details(
        {"viewId": 123, "viewName": "chart1", "header": ["name"], "details": [["alice"]]},
        user,
    )

    assert result == {"file": "chart1.xlsx", "status": "SUCCESS"}


@pytest.mark.asyncio
async def test_export_details_creates_chart_export_task_with_rows() -> None:
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))
    service.export_task_repo.create = AsyncMock(return_value=SimpleNamespace(id="task-2"))  # type: ignore[attr-defined]
    user = TokenUser(user_id=9, oid=1)

    payload: dict[str, object] = {
        "viewId": 123,
        "viewName": "chart1",
        "header": ["name", "value"],
        "details": [["alice", 1], ["bob", 2]],
    }

    result = await service.export_details(payload, user)

    assert result == {"file": "chart1.xlsx", "status": "SUCCESS"}
    service.export_task_repo.create.assert_awaited_once()  # type: ignore[attr-defined]
    created_payload = service.export_task_repo.create.await_args_list[0].args[0]  # type: ignore[attr-defined]
    assert created_payload["user_id"] == 9
    assert created_payload["export_from"] == 123
    assert created_payload["export_from_type"] == "chart"
    assert created_payload["params"] == {"data": [["name", "value"], ["alice", 1], ["bob", 2]]}


@pytest.mark.asyncio
async def test_export_details_returns_file_response_for_dataease_bi(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("DE_EXPORT_DIR", str(tmp_path))
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))
    user = TokenUser(user_id=7, oid=1)

    result = await service.export_details(
        {
            "viewId": 123,
            "viewName": "chart-bi",
            "dataEaseBi": True,
            "header": ["name", "value"],
            "details": [["alice", 1]],
        },
        user,
    )

    assert isinstance(result, FileResponse)
    file_response = cast(FileResponse, result)
    file_path = Path(cast(str, file_response.path))
    assert file_path.exists()
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows == [("name", "value"), ("alice", 1)]


def test_resolve_computed_expression_supports_double_encoded_origin_name() -> None:
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))
    expression = "[1715072798361]*[1715072798367]"
    double_encoded = base64.b64encode(base64.b64encode(expression.encode("utf-8"))).decode("utf-8")

    field = {
        "originName": double_encoded,
        "extField": 2,
    }
    all_fields = [
        {"id": "1715072798361", "originName": "单价"},
        {"id": "1715072798367", "originName": "销售数量"},
    ]

    resolved = service._resolve_computed_expression(field, all_fields)

    assert resolved == 'chart_source."单价"*chart_source."销售数量"'


def test_resolve_computed_expression_keeps_single_encoded_origin_name() -> None:
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))
    expression = "round(sum([7193537137675866112])/count([1715072798366])/100,2)"
    encoded = base64.b64encode(expression.encode("utf-8")).decode("utf-8")

    field = {
        "originName": encoded,
        "extField": 2,
    }
    all_fields = [
        {"id": "7193537137675866112", "originName": "销售金额"},
        {"id": "1715072798366", "originName": "账单流水号"},
    ]

    resolved = service._resolve_computed_expression(field, all_fields)

    assert resolved == 'round(sum(chart_source."销售金额")/count(chart_source."账单流水号")/100,2)'


def test_resolve_computed_expression_supports_nested_computed_references() -> None:
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))
    inner_expression = "[1715072798361]*[1715072798367]"
    outer_expression = "round(sum([7193537137675866112])/count([1715072798366])/100,2)"

    field = {
        "originName": base64.b64encode(base64.b64encode(outer_expression.encode("utf-8"))).decode("utf-8"),
        "extField": 2,
    }
    all_fields = [
        {
            "id": "7193537137675866112",
            "originName": base64.b64encode(base64.b64encode(inner_expression.encode("utf-8"))).decode("utf-8"),
            "extField": 2,
        },
        {"id": "1715072798361", "originName": "单价", "extField": 0},
        {"id": "1715072798367", "originName": "销售数量", "extField": 0},
        {"id": "1715072798366", "originName": "账单流水号", "extField": 0},
    ]

    resolved = service._resolve_computed_expression(field, all_fields)

    assert resolved == (
        'round(sum(chart_source."单价"*chart_source."销售数量")/'
        'count(chart_source."账单流水号")/100,2)'
    )


def test_resolve_computed_expression_supports_triple_encoded_origin_name() -> None:
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))
    expression = "[1715072798361]*[1715072798367]"
    triple_encoded = base64.b64encode(
        base64.b64encode(base64.b64encode(expression.encode("utf-8")))
    ).decode("utf-8")

    field = {
        "originName": triple_encoded,
        "extField": 2,
    }
    all_fields = [
        {"id": "1715072798361", "originName": "单价", "extField": 0},
        {"id": "1715072798367", "originName": "销售数量", "extField": 0},
    ]

    resolved = service._resolve_computed_expression(field, all_fields)

    assert resolved == 'chart_source."单价"*chart_source."销售数量"'


def test_build_chart_sql_does_not_wrap_aggregated_computed_expression_twice() -> None:
    service = ChartService(cast(AsyncSession, cast(Any, SimpleNamespace())))

    sql = service._build_chart_sql(
        'SELECT * FROM "demo_tea_order"',
        [],
        [
            {
                "id": "7193537244429291520",
                "dataeaseName": "f_39fd4542efb6a572",
                "name": "客单价",
                "summary": "sum",
                "extField": 2,
                "originName": base64.b64encode(
                    base64.b64encode(
                        b'round(sum([7193537137675866112])/count([1715072798366])/100,2)'
                    )
                ).decode("utf-8"),
            }
        ],
        1000,
        [
            {
                "id": "7193537137675866112",
                "originName": base64.b64encode(
                    base64.b64encode(b'[1715072798361]*[1715072798367]')
                ).decode("utf-8"),
                "extField": 2,
            },
            {"id": "1715072798361", "originName": "单价", "extField": 0},
            {"id": "1715072798367", "originName": "销售数量", "extField": 0},
            {"id": "1715072798366", "originName": "账单流水号", "extField": 0},
        ],
    )

    assert 'SUM(round(' not in sql
    assert (
        'round(sum(chart_source."单价"*chart_source."销售数量")/'
        'count(chart_source."账单流水号")/100,2) AS "f_39fd4542efb6a572"'
    ) in sql
