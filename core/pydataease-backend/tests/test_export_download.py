# pyright: reportMissingModuleSource=false
from __future__ import annotations

from pathlib import Path

import pytest
from fastapi import HTTPException, status
from openpyxl import load_workbook

from app.main import app
from app.services.export_service import get_export_service
from app.tasks.file_generator import generate_export_file
from tests.fixtures.auth_fixtures import _build_token


class FakeDownloadService:
    def __init__(self, result: dict[str, object] | None = None, *, missing: bool = False) -> None:
        self._result = result or {}
        self._missing = missing

    async def download(self, task_id: str) -> dict[str, object]:
        if self._missing:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Export task not found",
            )
        return {"id": task_id, **self._result}


@pytest.mark.asyncio
async def test_generate_export_file_with_data(tmp_path: Path) -> None:
    file_path, file_size = generate_export_file(
        task_id="task-1",
        file_name="report.xlsx",
        params={"data": [["name", "value"], ["alice", 1], ["bob", 2]]},
        export_dir=str(tmp_path),
    )

    assert Path(file_path).exists()
    assert file_size > 0

    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows == [("name", "value"), ("alice", 1), ("bob", 2)]


@pytest.mark.asyncio
async def test_generate_export_file_without_data(tmp_path: Path) -> None:
    file_path, file_size = generate_export_file(
        task_id="task-2",
        file_name="empty-report",
        params={},
        export_dir=str(tmp_path),
    )

    assert file_path.endswith("empty-report.xlsx")
    assert file_size > 0

    workbook = load_workbook(file_path)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows == [("No data",)]


@pytest.mark.asyncio
async def test_download_successful_task_returns_file(client, tmp_path: Path) -> None:
    export_file = tmp_path / "task-3.xlsx"
    export_file.write_bytes(b"file-body")

    service = FakeDownloadService(
        {
            "status": "SUCCESS",
            "path": str(export_file),
            "file_name": "task-3.xlsx",
        }
    )
    app.dependency_overrides[get_export_service] = lambda: service
    headers = {"X-DE-TOKEN": _build_token(uid=1, oid=1)}

    try:
        response = await client.get("/de2api/exportCenter/download/task-3", headers=headers)
    finally:
        _ = app.dependency_overrides.pop(get_export_service, None)

    assert response.status_code == 200
    assert response.content == b"file-body"
    assert response.headers["content-type"] == "application/octet-stream"
    assert "attachment; filename=\"task-3.xlsx\"" in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_download_non_existent_task_returns_404(client) -> None:
    service = FakeDownloadService(missing=True)
    app.dependency_overrides[get_export_service] = lambda: service
    headers = {"X-DE-TOKEN": _build_token(uid=1, oid=1)}

    try:
        response = await client.get("/de2api/exportCenter/download/missing-task", headers=headers)
    finally:
        _ = app.dependency_overrides.pop(get_export_service, None)

    assert response.status_code == 404
    assert response.json()["msg"] == "Export task not found"


@pytest.mark.asyncio
async def test_download_pending_task_returns_status_info(client) -> None:
    service = FakeDownloadService(
        {
            "status": "INITIATED",
            "msg": "Export file is not ready",
        }
    )
    app.dependency_overrides[get_export_service] = lambda: service
    headers = {"X-DE-TOKEN": _build_token(uid=1, oid=1)}

    try:
        response = await client.get("/de2api/exportCenter/download/pending-task", headers=headers)
    finally:
        _ = app.dependency_overrides.pop(get_export_service, None)

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["id"] == "pending-task"
    assert body["status"] == "INITIATED"
    assert body["msg"] == "Export file is not ready"
