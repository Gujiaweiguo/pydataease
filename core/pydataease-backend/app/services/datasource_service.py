from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import time
from collections.abc import Awaitable
from typing import Any, cast, final

import httpx
from openpyxl import load_workbook
from fastapi import Depends, HTTPException, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies.database import get_db
from app.models.datasource import CoreDatasource
from app.models.engine import CoreDeEngine
from app.repositories.datasource_repo import DatasourceRepository
from app.services.datasource_drivers import canonical_type, is_supported_type, open_connection
from app.utils.id_utils import _sid
from app.schemas.auth import TokenUser
from app.schemas.datasource import (
    JSONDict,
    JSONValue,
    DatasourceCreate,
    DatasourceFieldResponse,
    DatasourceResponse,
    DatasourceSimpleResponse,
    DatasourceTableResponse,
    DatasourceUpdate,
    DatasourceValidateRequest,
    DatasourceValidateResponse,
    EngineInfoResponse,
)

def _build_tree(nodes: list[dict[str, Any]], pid: str = "0") -> list[dict[str, Any]]:
    children: list[dict[str, Any]] = []
    for node in nodes:
        if node.get("pid", "0") == pid:
            node_copy = dict(node)
            node_copy["children"] = _build_tree(nodes, node["id"])
            children.append(node_copy)
    return children


@final
class DatasourceService:
    session: AsyncSession
    repository: DatasourceRepository

    def __init__(self, session: AsyncSession, repository: DatasourceRepository) -> None:
        self.session = session
        self.repository = repository

    async def query(self, keyword: str) -> list[dict[str, object]]:
        records = await self.repository.search(keyword)
        return [DatasourceResponse.model_validate(record).model_dump(by_alias=True) for record in records]

    async def tree(self) -> list[dict[str, Any]]:
        try:
            stmt = select(CoreDatasource).where(CoreDatasource.id != 0).order_by(CoreDatasource.name.asc(), CoreDatasource.update_time.desc())
            result = await self.session.execute(stmt)
            rows = result.scalars().all()
            flat = [
                {"id": _sid(row.id), "name": row.name or "", "pid": _sid(row.pid) if row.pid else "0",
                 "leaf": True, "weight": 9, "extraFlag": 0, "extraFlag1": 0}
                for row in rows
            ]
            children = _build_tree(flat, pid="0")
        except (AttributeError, TypeError):
            children = []
        root = {"id": "0", "name": "root", "pid": -1, "leaf": False,
                "weight": 7, "extraFlag": 0, "extraFlag1": 0, "children": children}
        return [root]

    async def save(self, payload: DatasourceCreate, user: TokenUser) -> dict[str, object]:
        if payload.id:
            update_payload = DatasourceUpdate(**payload.model_dump())
            return await self.update(update_payload, user)

        now = _timestamp_ms()
        status_value = await self._probe_status(payload.type, payload.configuration)
        pid_value = None if payload.pid == 0 else payload.pid
        configuration = self._persistable_configuration(payload.type, payload.configuration, payload.sync_setting)
        created = await self.repository.create(
            {
                "id": _new_identifier(),
                "name": payload.name.strip(),
                "description": payload.description,
                "type": payload.type,
                "pid": pid_value,
                "edit_type": payload.edit_type,
                "configuration": configuration,
                "create_time": now,
                "update_time": now,
                "update_by": user.user_id,
                "create_by": str(user.user_id),
                "status": status_value,
                "qrtz_instance": None,
                "task_status": "WaitingForExecution",
                "enable_data_fill": payload.enable_data_fill,
            }
        )
        return self._decorate_datasource_response(created)

    async def update(self, payload: DatasourceUpdate, user: TokenUser) -> dict[str, object]:
        existing = await self._get_entity(payload.id)
        merged_configuration = payload.configuration if payload.configuration is not None else existing.configuration
        merged_type = payload.type or existing.type
        merged_sync_setting = payload.sync_setting if payload.sync_setting is not None else self._extract_sync_setting(existing.configuration)
        merged_configuration = self._persistable_configuration(merged_type, merged_configuration, merged_sync_setting)

        raw_pid = payload.pid if payload.pid is not None else existing.pid
        pid_value = None if raw_pid == 0 else raw_pid
        updated = await self.repository.update(
            existing,
            {
                "name": payload.name.strip() if payload.name is not None else existing.name,
                "description": payload.description if payload.description is not None else existing.description,
                "type": merged_type,
                "pid": pid_value,
                "edit_type": payload.edit_type if payload.edit_type is not None else existing.edit_type,
                "configuration": merged_configuration,
                "update_time": _timestamp_ms(),
                "update_by": user.user_id,
                "status": await self._probe_status(merged_type, merged_configuration),
                "enable_data_fill": payload.enable_data_fill if payload.enable_data_fill is not None else existing.enable_data_fill,
            },
        )
        return self._decorate_datasource_response(updated)

    async def delete(self, datasource_id: int) -> None:
        entity = await self._get_entity(datasource_id)
        if await self.check_in_use(datasource_id):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Datasource is still referenced by datasets and cannot be deleted",
            )
        await self.repository.delete(entity)

    async def validate(self, payload: DatasourceValidateRequest) -> DatasourceValidateResponse:
        if self._is_file_type(payload.type):
            return DatasourceValidateResponse(success=True, message="Connection successful", datasource_type=payload.type)
        self._ensure_supported_type(payload.type)
        await self._validate_connection(payload.type, _as_config_dict(payload.configuration))
        return DatasourceValidateResponse(success=True, message="Connection successful", datasource_type=payload.type)

    async def get_schemas_from_config(self, configuration: JSONDict, ds_type: str = "postgresql") -> list[str]:
        if not is_supported_type(ds_type):
            return []
        connection = await self._open_connection(configuration, ds_type)
        try:
            rows = await connection.fetch(
                "SELECT schema_name FROM information_schema.schemata ORDER BY schema_name"
            )
        finally:
            await connection.close()
        return [row["schema_name"] for row in rows]

    async def get_tables(self, datasource_id: int) -> list[DatasourceTableResponse]:
        datasource = await self._get_entity(datasource_id)
        if self._is_file_type(datasource.type):
            sheets = self._excel_sheets(datasource.configuration)
            return [
                DatasourceTableResponse(
                    name=str(sheet.get("deTableName") or sheet.get("tableName") or ""),
                    table_name=str(sheet.get("tableName") or sheet.get("deTableName") or ""),
                    schema_name="public",
                    last_update_time=datasource.update_time,
                    status="Completed",
                )
                for sheet in sheets
            ]
        ds_type = canonical_type(datasource.type)
        configuration = _as_config_dict(datasource.configuration)
        connection = await self._open_connection(configuration, datasource.type)
        schema = self._schema_name(configuration, ds_type)
        try:
            if ds_type == "postgresql":
                rows = await connection.fetch(
                    """
                    SELECT c.relname AS table_name,
                           obj_description(c.oid) AS remark
                    FROM pg_class c
                    JOIN pg_namespace n ON n.oid = c.relnamespace
                    WHERE n.nspname = $1 AND c.relkind = 'r'
                    ORDER BY c.relname
                    """,
                    schema,
                )
            else:
                # MySQL and others: use information_schema
                rows = await connection.fetch(
                    """
                    SELECT table_name,
                           table_comment AS remark
                    FROM information_schema.tables
                    WHERE table_schema = $1 AND table_type = 'BASE TABLE'
                    ORDER BY table_name
                    """,
                    schema,
                )
        finally:
            await connection.close()
        return [
            DatasourceTableResponse(
                name=row["remark"] or row["table_name"],
                table_name=row["table_name"],
                schema_name=schema,
                last_update_time=datasource.update_time,
                status="Completed",
            )
            for row in rows
        ]

    async def get_fields(self, datasource_id: int, table_name: str) -> list[DatasourceFieldResponse]:
        datasource = await self._get_entity(datasource_id)
        if self._is_file_type(datasource.type):
            for sheet in self._excel_sheets(datasource.configuration):
                current_name = str(sheet.get("deTableName") or sheet.get("tableName") or "")
                if current_name == table_name or str(sheet.get("tableName") or "") == table_name:
                    return [self._excel_datasource_field(field) for field in cast(list[dict[str, object]], sheet.get("fields", []))]
            return []
        ds_type = canonical_type(datasource.type)
        configuration = _as_config_dict(datasource.configuration)
        connection = await self._open_connection(configuration, datasource.type)
        schema = self._schema_name(configuration, ds_type)
        try:
            rows = await connection.fetch(
                """
                SELECT column_name, data_type, is_nullable
                FROM information_schema.columns
                WHERE table_schema = $1 AND table_name = $2
                ORDER BY ordinal_position
                """,
                schema,
                table_name,
            )
        finally:
            await connection.close()
        return [self._db_datasource_field(row) for row in rows]

    async def get_engine_info(self) -> EngineInfoResponse:
        result = await self.session.execute(select(CoreDeEngine).limit(1))
        engine = result.scalars().first()
        if engine is None:
            return EngineInfoResponse(configured=False, type=None, status=None, name=None)
        return EngineInfoResponse(
            configured=True,
            type=engine.type,
            status=engine.status,
            name=engine.name,
        )

    async def latest_use(self) -> list[str]:
        try:
            stmt = (
                select(CoreDatasource.type)
                .where(CoreDatasource.type != "folder")
                .order_by(CoreDatasource.create_time.desc())
                .limit(5)
            )
            result = await self.session.execute(stmt)
            types = list(set(row[0] for row in result.all() if row[0]))
            return types[:5]
        except (AttributeError, TypeError):
            return []

    async def move(self, datasource_id: int, pid: int) -> None:
        entity = await self._get_entity(datasource_id)
        pid_value = None if pid == 0 else pid
        await self.repository.update(entity, {"pid": pid_value})

    async def rename(self, datasource_id: int, name: str) -> None:
        entity = await self._get_entity(datasource_id)
        await self.repository.update(entity, {"name": name.strip(), "update_time": _timestamp_ms()})

    async def create_folder(self, name: str, pid: int, user: TokenUser) -> dict[str, object]:
        now = _timestamp_ms()
        pid_value = None if pid == 0 else pid
        created = await self.repository.create(
            {
                "id": _new_identifier(),
                "name": name.strip(),
                "type": "folder",
                "pid": pid_value,
                "configuration": {},
                "description": None,
                "edit_type": None,
                "create_time": now,
                "update_time": now,
                "update_by": user.user_id,
                "create_by": str(user.user_id),
                "status": "Success",
                "qrtz_instance": None,
                "task_status": "WaitingForExecution",
                "enable_data_fill": None,
            }
        )
        return self._decorate_datasource_response(created)

    async def get_full(self, datasource_id: int) -> dict[str, object]:
        entity = await self._get_entity(datasource_id)
        return self._decorate_datasource_response(entity)

    async def get_hide_pw(self, datasource_id: int) -> dict[str, object]:
        entity = await self._get_entity(datasource_id)
        data = self._decorate_datasource_response(entity)
        config = data.get("configuration")
        if config and isinstance(config, dict):
            _mask_passwords(config)
        return data

    async def get_simple(self, datasource_id: int) -> dict[str, object]:
        entity = await self._get_entity(datasource_id)
        return DatasourceSimpleResponse.model_validate(entity).model_dump(by_alias=True)

    async def validate_by_id(self, datasource_id: int) -> dict[str, str]:
        entity = await self._get_entity(datasource_id)
        config = entity.configuration if isinstance(entity.configuration, dict) else {}
        ds_type = entity.type
        status_value = await self._probe_status(ds_type, config)
        return {"status": status_value, "type": ds_type}

    async def get_table_status(self, datasource_id: int) -> list[dict[str, object]]:
        datasource = await self._get_entity(datasource_id)
        if self._is_file_type(datasource.type):
            now = datasource.update_time
            return [
                {
                    "tableName": str(sheet.get("deTableName") or sheet.get("tableName") or ""),
                    "lastUpdateTime": now,
                    "status": "Completed",
                }
                for sheet in self._excel_sheets(datasource.configuration)
            ]
        tables = await self.get_tables(datasource_id)
        return [
            {
                "tableName": table.name,
                "lastUpdateTime": datasource.update_time,
                "status": "Completed",
            }
            for table in tables
        ]

    async def check_repeat(self, payload: dict[str, object]) -> bool:
        ds_type = str(payload.get("type") or "")
        if ds_type == "folder" or "API" in ds_type or "Excel" in ds_type or ds_type in {"es"}:
            return False
        configuration = payload.get("configuration")
        if not isinstance(configuration, dict):
            return False
        host = configuration.get("host")
        port = configuration.get("port")
        database = configuration.get("database") or configuration.get("dataBase")
        schema = configuration.get("schema")
        stmt = select(CoreDatasource).where(CoreDatasource.type == ds_type)
        result = await self.session.execute(stmt)
        for row in result.scalars().all():
            if payload.get("id") and str(row.id) == str(payload.get("id")):
                continue
            row_cfg = row.configuration if isinstance(row.configuration, dict) else {}
            if (
                row_cfg.get("host") == host
                and str(row_cfg.get("port")) == str(port)
                and (row_cfg.get("database") or row_cfg.get("dataBase")) == database
                and (schema is None or row_cfg.get("schema") == schema)
            ):
                return True
        return False

    async def preview_data(self, payload: dict[str, object]) -> dict[str, object]:
        raw_datasource_id = payload.get("id") or payload.get("datasourceId") or 0
        datasource_id = int(str(raw_datasource_id))
        table_name = str(payload.get("table") or payload.get("tableName") or "")
        datasource = await self._get_entity(datasource_id)
        if self._is_file_type(datasource.type):
            for sheet in self._excel_sheets(datasource.configuration):
                current_name = str(sheet.get("deTableName") or sheet.get("tableName") or "")
                if current_name == table_name or str(sheet.get("tableName") or "") == table_name:
                    fields = [self._excel_preview_field(field) for field in cast(list[dict[str, object]], sheet.get("fields", []))]
                    data_rows = sheet.get("jsonArray", [])
                    # The frontend (ExcelDetail.vue:356-357) clears jsonArray
                    # before saving.  Re-parse from the stored raw file content
                    # when the in-config data is empty.
                    if not data_rows and sheet.get("_rawFileBase64"):
                        data_rows = self._reparse_excel_rows(
                            sheet, str(sheet.get("_rawFileBase64"))
                        )
                    return {"data": {"fields": fields, "data": data_rows}}
            return {"data": {"fields": [], "data": []}}
        fields = await self.get_fields(datasource_id, table_name)
        datasource_cfg = _as_config_dict(datasource.configuration)
        connection = await self._open_connection(datasource_cfg, datasource.type)
        schema = self._schema_name(datasource_cfg, canonical_type(datasource.type))
        # BUG-002 fix: Validate identifiers to prevent SQL injection
        table_name = self._validate_identifier(table_name, "table")
        schema = self._validate_identifier(schema, "schema")
        try:
            rows = await connection.fetch(
                f'SELECT * FROM "{schema}"."{table_name}" LIMIT 10' if canonical_type(datasource.type) == "postgresql" else f"SELECT * FROM `{table_name}` LIMIT 10"
            )
        finally:
            await connection.close()
        # Convert asyncpg Records to dicts keyed by column name so the
        # frontend el-table-v2 dataKey lookup works correctly.
        row_dicts = [dict(row) for row in rows]
        preview_fields = [
            {"name": field.name, "originName": field.name, "deType": self._field_de_type(field.data_type), "fieldType": field.data_type}
            for field in fields
        ]
        return {"data": {"fields": preview_fields, "data": row_dicts}}

    async def sync_api_table(self, payload: dict[str, object]) -> None:
        return None

    async def sync_api_datasource(self, payload: dict[str, object]) -> None:
        return None

    async def list_sync_record(self, datasource_id: int, page: int, limit: int) -> dict[str, object]:
        return {"records": [], "total": 0, "page": page, "pageSize": limit}

    async def check_api_datasource(self, payload: dict[str, object]) -> dict[str, object]:
        return {"jsonFields": [], "fields": [], "data": []}

    async def check_in_use(self, datasource_id: int) -> bool:
        from app.models.dataset import CoreDatasetTable

        stmt = select(func.count()).select_from(CoreDatasetTable).where(
            CoreDatasetTable.datasource_id == datasource_id
        )
        result = await self.session.execute(stmt)
        count = result.scalar() or 0
        return count > 0

    async def _get_entity(self, datasource_id: int) -> CoreDatasource:
        entity = await self.repository.get_by_id(datasource_id)
        if entity is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Datasource not found")
        return entity

    def _ensure_supported_type(self, datasource_type: str) -> None:
        if self._is_file_type(datasource_type):
            return
        if not is_supported_type(datasource_type):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Datasource type '{datasource_type}' is not supported. "
                    "Supported: PostgreSQL (pg/postgres/postgresql), MySQL (mysql/mariadb)."
                ),
            )

    async def _probe_status(self, datasource_type: str, configuration: JSONValue) -> str:
        try:
            if self._is_file_type(datasource_type):
                return "Success"
            self._ensure_supported_type(datasource_type)
            await self._validate_connection(datasource_type, _as_config_dict(configuration))
        except Exception:
            return "Error"
        return "Success"

    async def _validate_connection(self, ds_type: str, configuration: JSONDict) -> None:
        connection = await self._open_connection(configuration, ds_type)
        try:
            _ = await connection.fetchval("SELECT 1")
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Connection failed: {exc}") from exc
        finally:
            await connection.close()

    async def _open_connection(self, configuration: JSONDict, ds_type: str | None = None) -> Any:
        type_str = ds_type or "postgresql"
        try:
            return await open_connection(type_str, configuration)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Connection failed: {exc}") from exc

    @staticmethod
    def _validate_identifier(name: str, field: str = "table") -> str:
        """Validate a SQL identifier (table name, schema) to prevent injection."""
        import re

        if not re.match(r"^[a-zA-Z_][a-zA-Z0-9_.]*$", name):
            raise HTTPException(status_code=400, detail=f"Invalid {field} name")
        return name

    @staticmethod
    def _schema_name(configuration: JSONDict, ds_type: str = "postgresql") -> str:
        schema = configuration.get("schema") or configuration.get("currentSchema")
        if schema:
            return str(schema)
        if ds_type == "mysql":
            database = configuration.get("database") or configuration.get("dataBase") or "mysql"
            return str(database)
        schema = "public"
        return str(schema)

    async def upload_file(self, file: object, id: str | None = None, edit_type: str | None = None) -> dict[str, object]:
        filename = str(getattr(file, "filename", "") or "excel.xlsx")
        content = await cast(Any, file).read()
        return self._parse_uploaded_file(content, filename)

    async def load_remote_file(self, payload: dict[str, object]) -> dict[str, object]:
        url = str(payload.get("url") or "")
        # BUG-006 fix: Validate URL scheme to prevent SSRF
        from urllib.parse import urlparse

        parsed = urlparse(url)
        if parsed.scheme not in ("http", "https"):
            raise HTTPException(status_code=400, detail="Only HTTP/HTTPS URLs are supported")
        if not parsed.hostname:
            raise HTTPException(status_code=400, detail="Invalid URL: missing hostname")

        username = self._decode_b64_string(payload.get("userName"))
        password = self._decode_b64_string(payload.get("passwd"))
        result = cast(object, self._download_remote_file(url, username, password))
        content: bytes
        filename: str
        if hasattr(result, "__await__"):
            content, filename = await cast(Awaitable[tuple[bytes, str]], result)
        else:
            content, filename = cast(tuple[bytes, str], result)
        return self._parse_uploaded_file(content, filename)

    @staticmethod
    async def _download_remote_file(url: str, username: str, password: str) -> tuple[bytes, str]:
        auth = (username, password) if username or password else None
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            response = await client.get(url, auth=auth)
            response.raise_for_status()
            filename = url.rstrip("/").split("/")[-1] or "remote.xlsx"
            return response.content, filename

    def _parse_uploaded_file(self, content: bytes, filename: str) -> dict[str, object]:
        lower = filename.lower()
        if lower.endswith(".csv"):
            sheets = [self._parse_csv_sheet(content, filename)]
        else:
            sheets = self._parse_excel_sheets(content)
        # Embed raw file content as base64 so the frontend includes it in the
        # save payload.  The frontend clears *data* and *jsonArray* before
        # saving (ExcelDetail.vue:356-357) but leaves all other keys intact.
        # When preview_data sees an empty jsonArray it can re-parse from the
        # stored file content.
        _raw = base64.b64encode(content).decode()
        for s in sheets:
            s["_rawFileBase64"] = _raw
        return {"excelLabel": filename, "sheets": sheets, "fileName": filename}

    def _parse_excel_sheets(self, content: bytes) -> list[dict[str, object]]:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheets: list[dict[str, object]] = []
        for index, worksheet in enumerate(workbook.worksheets, start=1):
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(cell or "") for cell in rows[0]]
            json_rows = cast(list[dict[str, object]], [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)])
            fields = [self._excel_field(header, json_rows) for header in headers]
            table_name = worksheet.title or f"sheet_{index}"
            sheets.append(
                {
                    "sheet": True,
                    "sheetId": str(index),
                    "sheetExcelId": str(index),
                    "tableName": table_name,
                    "deTableName": self._excel_table_name(table_name),
                    "excelLabel": table_name,
                    "fields": fields,
                    "jsonArray": json_rows,
                    "fieldsMd5": self._fields_md5(fields),
                    "newSheet": True,
                }
            )
        return sheets

    def _parse_csv_sheet(self, content: bytes, filename: str) -> dict[str, object]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = cast(list[dict[str, object]], list(reader))
        headers = reader.fieldnames or []
        fields = [self._excel_field(header, rows) for header in headers]
        table_name = filename.rsplit(".", 1)[0]
        return {
            "sheet": True,
            "sheetId": "1",
            "sheetExcelId": "1",
            "tableName": table_name,
            "deTableName": self._excel_table_name(table_name),
            "excelLabel": filename,
            "fields": fields,
            "jsonArray": rows,
            "fieldsMd5": self._fields_md5(fields),
            "newSheet": True,
        }

    def _reparse_excel_rows(self, sheet: dict[str, object], raw_b64: str) -> list[dict[str, object]]:
        """Re-parse row data from the raw file content stored in ``_rawFileBase64``.

        The frontend clears *jsonArray* before saving (``ExcelDetail.vue:356-357``).
        We stash the original file bytes as base64 so that ``preview_data`` can
        materialise the rows on demand instead of requiring a live file on disk
        (the Java backend's approach).
        """
        try:
            content = base64.b64decode(raw_b64)
        except Exception:
            return []
        table_name = str(sheet.get("tableName") or "")
        target = str(sheet.get("deTableName") or table_name)
        # Try Excel first (openpyxl); fall back to CSV if the content is not
        # a valid ZIP/XLSX archive.
        try:
            parsed_sheets = self._parse_excel_sheets(content)
        except Exception:
            parsed_sheets = []
        if parsed_sheets:
            for ps in parsed_sheets:
                pn = str(ps.get("deTableName") or ps.get("tableName") or "")
                if pn == target or str(ps.get("tableName") or "") == table_name:
                    return cast(list[dict[str, object]], ps.get("jsonArray", []))
        # CSV fallback — single sheet, just re-parse.
        try:
            csv_name = table_name if table_name.lower().endswith(".csv") else table_name + ".csv"
            parsed = self._parse_csv_sheet(content, csv_name)
            return cast(list[dict[str, object]], parsed.get("jsonArray", []))
        except Exception:
            return []

    def _excel_sheets(self, configuration: object) -> list[dict[str, object]]:
        if isinstance(configuration, list):
            return cast(list[dict[str, object]], configuration)
        if isinstance(configuration, dict):
            sheets = configuration.get("sheets")
            if isinstance(sheets, list):
                return cast(list[dict[str, object]], sheets)
        return []

    def _excel_field(self, header: str, rows: list[dict[str, object]] | list[dict[str, Any]]) -> dict[str, object]:
        values = [row.get(header) for row in rows if header in row]
        field_type, de_type = self._infer_excel_field(values)
        return {
            "name": header,
            "originName": header,
            "fieldType": field_type,
            "deType": de_type,
            "deExtractType": de_type,
            "checked": True,
            "primaryKey": False,
            "length": 255 if de_type == 0 else None,
        }

    def _excel_preview_field(self, field: dict[str, object]) -> dict[str, object]:
        return {
            "name": field.get("name") or field.get("originName"),
            "originName": field.get("originName") or field.get("name"),
            "deType": field.get("deType") if field.get("deType") is not None else field.get("deExtractType", 0),
            "fieldType": field.get("fieldType", "TEXT"),
        }

    def _infer_excel_field(self, values: list[object]) -> tuple[str, int]:
        concrete = [value for value in values if value not in (None, "")]
        if not concrete:
            return "TEXT", 0
        if all(isinstance(value, bool) for value in concrete):
            return "TEXT", 0
        if all(isinstance(value, int) for value in concrete):
            return "LONG", 2
        if all(isinstance(value, int | float) for value in concrete):
            return "DOUBLE", 3
        return "TEXT", 0

    def _fields_md5(self, fields: list[dict[str, object]]) -> str:
        normalized = json.dumps([{k: field.get(k) for k in ("originName", "fieldType", "deExtractType")} for field in fields], sort_keys=True)
        return hashlib.md5(normalized.encode()).hexdigest()

    def _excel_table_name(self, table_name: str) -> str:
        safe = "".join(ch if ch.isalnum() else "_" for ch in table_name.lower()).strip("_") or "sheet"
        return f"excel_{safe}_{hashlib.md5(table_name.encode()).hexdigest()[:10]}"

    def _field_de_type(self, data_type: str) -> int:
        upper = data_type.upper()
        if upper in {"LONG", "INTEGER", "BIGINT", "INT"}:
            return 2
        if upper in {"DOUBLE", "FLOAT", "DECIMAL", "NUMERIC", "REAL"}:
            return 3
        if upper in {"DATETIME", "TIMESTAMP", "DATE"}:
            return 1
        return 0

    @staticmethod
    def _is_file_type(datasource_type: str) -> bool:
        return datasource_type in {"Excel", "ExcelRemote"}

    @staticmethod
    def _decode_b64_string(value: object) -> str:
        if not isinstance(value, str) or value == "":
            return ""
        try:
            return base64.b64decode(value).decode()
        except Exception:
            return value

    def _db_datasource_field(self, row: dict[str, object]) -> DatasourceFieldResponse:
        data_type = str(row.get("data_type") or "TEXT")
        name = str(row.get("column_name") or "")
        return DatasourceFieldResponse(
            name=name,
            origin_name=name,
            data_type=data_type,
            de_type=self._field_de_type(data_type),
            type=data_type,
            nullable=str(row.get("is_nullable") or "YES") == "YES",
        )

    def _excel_datasource_field(self, field: dict[str, object]) -> DatasourceFieldResponse:
        name = str(field.get("originName") or field.get("name") or "")
        data_type = str(field.get("fieldType") or "TEXT")
        return DatasourceFieldResponse(
            name=name,
            origin_name=name,
            data_type=data_type,
            de_type=int(str(field.get("deType", self._field_de_type(data_type)))),
            type=data_type,
            nullable=not bool(field.get("primaryKey", False)),
        )

    def _decorate_datasource_response(self, entity: CoreDatasource) -> dict[str, object]:
        data = DatasourceResponse.model_validate(entity).model_dump(by_alias=True)
        if entity.type.startswith("API") and isinstance(entity.configuration, list):
            api_items, param_items = self._split_api_configuration(entity.configuration)
            data["apiConfigurationStr"] = api_items
            data["paramsStr"] = param_items
            data["configuration"] = {}
            if data.get("syncSetting") is None:
                data["syncSetting"] = self._default_sync_setting()
            return data
        if self._is_file_type(entity.type):
            sheets = self._excel_sheets(entity.configuration)
            data["fileName"] = data.get("fileName") or f"{entity.name}.xlsx"
            data["size"] = data.get("size") or len(json.dumps(sheets))
            if entity.type == "ExcelRemote":
                if isinstance(entity.configuration, dict):
                    config = cast(dict[str, object], dict(entity.configuration))
                    sync_setting = config.get("syncSetting") if isinstance(config.get("syncSetting"), dict) else None
                    if sync_setting is not None:
                        data["syncSetting"] = sync_setting
            return data
        return data

    def _persistable_configuration(
        self,
        datasource_type: str,
        configuration: JSONValue,
        sync_setting: JSONDict | None,
    ) -> JSONValue:
        if datasource_type == "ExcelRemote":
            if isinstance(configuration, list):
                data: JSONDict = {"sheets": configuration}
            elif isinstance(configuration, dict):
                data = dict(configuration)
            else:
                return configuration
            if sync_setting is not None:
                data["syncSetting"] = sync_setting
            return data
        return configuration

    def _extract_sync_setting(self, configuration: JSONValue) -> JSONDict | None:
        if isinstance(configuration, dict):
            sync_setting = configuration.get("syncSetting")
            if isinstance(sync_setting, dict):
                return cast(JSONDict, sync_setting)
        return None

    def _split_api_configuration(self, configuration: list[object]) -> tuple[list[object], list[object]]:
        api_items: list[object] = []
        param_items: list[object] = []
        for item in configuration:
            if isinstance(item, dict) and item.get("type") == "params":
                param_items.append(item)
            else:
                api_items.append(item)
        return api_items, param_items

    def _default_sync_setting(self) -> JSONDict:
        return {
            "updateType": "all_scope",
            "syncRate": "SIMPLE_CRON",
            "simpleCronValue": "1",
            "simpleCronType": "minute",
            "startTime": 0,
            "endTime": 0,
            "endLimit": "0",
            "cron": "0 0/1 * * * ? *",
        }


async def get_datasource_service(session: AsyncSession = Depends(get_db)) -> DatasourceService:
    return DatasourceService(session=session, repository=DatasourceRepository(session))

def _as_config_dict(configuration: object) -> JSONDict:
    if not isinstance(configuration, dict):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Datasource configuration must be an object")
    return cast(JSONDict, configuration)


def _timestamp_ms() -> int:
    return int(time.time() * 1000)


def _new_identifier() -> int:
    return time.time_ns()


def _mask_passwords(config: dict[str, object]) -> None:
    for key in list(config.keys()):
        if key.lower() in ("password", "pwd", "pass", "accesskey", "secretkey"):
            config[key] = "******"
        elif isinstance(config[key], dict):
            _mask_passwords(cast(dict[str, object], config[key]))
